from datetime import datetime

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .forms import CategoryForm, ProductEditForm, ProductForm
from .services import CategoryService, ProductService

# =========================
# PRODUTOS
# =========================


class ProductListCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Lista produtos com filtros"""

    def get_permission_required(self):
        return ["products.view_product"]

    def get(self, request):
        products = ProductService.get_all()

        # Filtro por categoria
        category_id = request.GET.get("category")
        if category_id and category_id.strip() and category_id != "None":
            products = products.filter(category_id=category_id)

        # Filtro por lote
        batch = request.GET.get("batch")
        if batch and batch.strip() and batch != "None":
            products = products.filter(batch__icontains=batch)

        # Filtro por data de vencimento
        expiration_date = request.GET.get("expiration_date")
        if expiration_date and expiration_date.strip() and expiration_date != "None":
            products = products.filter(expiration_date=expiration_date)

        # Obter todas as categorias para o dropdown
        categories = CategoryService.get_all()

        return render(
            request,
            "products/list.html",
            {
                "products": products,
                "categories": categories,
                "active_page": "products",
                "filter_category": category_id,
                "filter_batch": batch,
                "filter_date": expiration_date,
            },
        )


class ProductExportView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Exporta produtos para Excel"""

    def get_permission_required(self):
        return ["products.view_product"]

    def post(self, request):
        try:
            products = ProductService.get_all()

            # Aplicar os mesmos filtros
            category_id = request.POST.get("category")
            if category_id and category_id.strip() and category_id != "None":
                products = products.filter(category_id=category_id)

            batch = request.POST.get("batch")
            if batch and batch.strip() and batch != "None":
                products = products.filter(batch__icontains=batch)

            expiration_date = request.POST.get("expiration_date")
            if (
                expiration_date
                and expiration_date.strip()
                and expiration_date != "None"
            ):
                products = products.filter(expiration_date=expiration_date)

            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Produtos"

            # Estilos
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            header_fill = PatternFill(
                start_color="FF8C00", end_color="FF8C00", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            center_alignment = Alignment(horizontal="center", vertical="center")

            # Headers
            headers = [
                "ID",
                "Nome",
                "Categoria",
                "Lote",
                "Quantidade",
                "Preço (R$)",
                "Vencimento",
                "Status",
            ]
            ws.append(headers)

            # Aplicar estilo ao header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

            # Dados
            for product in products:
                # Determinar status
                if product.status == "ok":
                    status_text = "OK"
                elif product.status == "baixo":
                    status_text = "Abaixo do Estoque"
                elif product.status == "proximo_vencimento":
                    status_text = "Próximo Vencimento"
                else:
                    status_text = "-"

                row = [
                    product.id,
                    product.name,
                    product.category.name,
                    product.batch or "-",
                    product.quantity,
                    f"R$ {product.price:.2f}",
                    product.expiration_date.strftime("%d/%m/%Y"),
                    status_text,
                ]
                ws.append(row)

                # Aplicar border a todas as células
                for cell in ws[ws.max_row]:
                    cell.border = border
                    if ws.max_row > 1:
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = center_alignment

            # Ajustar largura das colunas
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 15
            ws.column_dimensions["G"].width = 15
            ws.column_dimensions["H"].width = 20

            # Preparar resposta
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = (
                f'attachment; filename="produtos_{datetime.now().strftime("%d_%m_%Y_%H%M%S")}.xlsx"'
            )

            wb.save(response)
            return response

        except Exception as e:
            print(f"\n=== ERRO NA EXPORTAÇÃO ===")
            print(f"Erro: {str(e)}")
            import traceback

            traceback.print_exc()
            print(f"=======================\n")
            messages.error(request, f"Erro ao exportar: {str(e)}")
            return redirect("product-list-create")


class ProductCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Cria novo produto"""

    permission_required = "products.add_product"

    def get(self, request):
        form = ProductForm()
        return render(
            request,
            "products/create.html",
            {"form": form, "active_page": "products"},
        )

    def post(self, request):
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                product = ProductService.create_product(request)
                if product:
                    messages.success(request, "O produto foi criado com sucesso!")
                    return redirect("product-list-create")
                else:
                    messages.error(request, "Erro ao criar o produto!")
            except Exception as e:
                messages.error(request, f"Erro ao salvar o produto: {str(e)}")
        else:
            messages.error(request, "Verifique se os dados inseridos estão corretos!")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        return render(
            request,
            "products/create.html",
            {"form": form, "active_page": "products"},
        )


class ProductUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Edita um produto existente"""

    def get_permission_required(self):
        return ["products.change_product"]

    def get(self, request, id_product):
        product = ProductService.get_product_by_id(id_product)
        if not product:
            messages.error(request, "Produto não encontrado.")
            return redirect("product-list-create")

        form = ProductEditForm(instance=product)
        return render(
            request,
            "products/edit.html",
            {"form": form, "product": product, "active_page": "products"},
        )

    def post(self, request, id_product):
        product = ProductService.get_product_by_id(id_product)
        if not product:
            messages.error(request, "Produto não encontrado.")
            return redirect("product-list-create")

        form = ProductEditForm(request.POST, instance=product)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "O produto foi modificado com sucesso!")
                return redirect("product-list-create")
            except Exception as e:
                messages.error(request, f"Erro ao atualizar o produto: {str(e)}")
        else:
            messages.error(request, "Os dados enviados não são válidos.")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        return render(
            request,
            "products/edit.html",
            {"form": form, "product": product, "active_page": "products"},
        )


class ProductDeleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Deleta um produto"""

    permission_required = "products.delete_product"

    def post(self, request, id_product):
        try:
            deleted = ProductService.delete_product_by_id(id_product)
            if deleted:
                messages.success(request, "Produto deletado com sucesso!")
            else:
                messages.error(request, "Produto não encontrado.")
        except Exception as e:
            messages.error(request, f"Erro ao excluir o produto: {str(e)}")

        return redirect("product-list-create")


# =========================
# CATEGORIAS
# =========================


class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Lista todas as categorias"""

    def get_permission_required(self):
        return ["products.view_category"]

    def get(self, request):
        categories = CategoryService.get_all()
        return render(
            request,
            "products/categories/index.html",
            {
                "categories": categories,
                "active_page": "categories",
            },
        )


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Cria nova categoria"""

    permission_required = "products.add_category"

    def get(self, request):
        form = CategoryForm()
        return render(
            request,
            "products/categories/create.html",
            {
                "form": form,
                "active_page": "categories",
            },
        )

    def post(self, request):
        form = CategoryForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "A categoria foi criada com sucesso!")
                return redirect("category-list-create")
            except Exception as e:
                messages.error(request, f"Erro ao criar categoria: {str(e)}")
        else:
            messages.error(request, "Verifique se os dados inseridos estão corretos!")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        return render(
            request,
            "products/categories/create.html",
            {
                "form": form,
                "active_page": "categories",
            },
        )


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Edita uma categoria"""

    def get_permission_required(self):
        return ["products.change_category"]

    def get(self, request, id_category):
        category = CategoryService.get_category_by_id(id_category)
        if not category:
            messages.error(request, "Categoria não encontrada.")
            return redirect("category-list-create")

        form = CategoryForm(instance=category)
        return render(
            request,
            "products/categories/edit.html",
            {
                "form": form,
                "category": category,
                "active_page": "categories",
            },
        )

    def post(self, request, id_category):
        category = CategoryService.get_category_by_id(id_category)
        if not category:
            messages.error(request, "Categoria não encontrada.")
            return redirect("category-list-create")

        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Categoria atualizada com sucesso!")
                return redirect("category-list-create")
            except Exception as e:
                messages.error(request, f"Erro ao atualizar categoria: {str(e)}")
        else:
            messages.error(request, "Os dados enviados não são válidos.")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        return render(
            request,
            "products/categories/edit.html",
            {
                "form": form,
                "category": category,
                "active_page": "categories",
            },
        )


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Deleta uma categoria"""

    permission_required = "products.delete_category"

    def post(self, request, id_category):
        try:
            deleted = CategoryService.delete_category_by_id(id_category)
            if deleted:
                messages.success(request, "Categoria deletada com sucesso!")
            else:
                messages.error(request, "Categoria não encontrada.")
        except Exception as e:
            messages.error(request, f"Erro ao deletar categoria: {str(e)}")

        return redirect("category-list-create")
